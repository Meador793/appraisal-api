"""
Excel export of the analysis.

Mirrors the original regression notebook's Step 17 -- one workbook, one sheet
per result table, suitable for an appraisal work file.

A note on formulas: there are none here, deliberately. The guidance for a
financial model is to write formulas so the sheet recalculates when inputs
change, and that is right for a model. This is not a model; it is a record of
what a trained gradient-boosted ensemble produced. An XGBoost marginal effect
cannot be expressed as an Excel formula, so writing one would either be a lie
about how the number was derived or a hardcoded value dressed up as a
calculation. The Basis sheet says so explicitly, and names the source file and
model version so the reader can reproduce it.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name=FONT, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=12, color="1F3864")
NOTE_FONT = Font(name=FONT, size=9, italic=True, color="555555")
THIN = Side(style="thin", color="D5DBE5")

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.00"%"'
NUM = '#,##0'
DEC4 = '0.0000'


def _write(ws, df: pd.DataFrame, start_row: int = 1, formats: dict | None = None) -> int:
    """Write a DataFrame with a styled header. Returns the next free row."""
    formats = formats or {}
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=start_row, column=j, value=str(col))
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            if isinstance(v, (pd.Timestamp,)):
                v = str(v.date())
            elif hasattr(v, "item"):
                v = v.item()
            c = ws.cell(row=i, column=j, value=v)
            c.font = BODY_FONT
            c.border = Border(bottom=THIN)
            if col in formats:
                c.number_format = formats[col]

    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)) + 4,
                    *(len(str(v)) + 3 for v in df[col].astype(str).head(60))) if len(df) else len(str(col)) + 4
        ws.column_dimensions[get_column_letter(j)].width = min(max(width, 10), 42)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(df) + 2


def _note(ws, row: int, text: str, col: int = 1) -> int:
    c = ws.cell(row=row, column=col, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def build_workbook(meta: dict, grid: pd.DataFrame, pct: pd.DataFrame,
                   loc: pd.DataFrame, importance: pd.DataFrame,
                   X: pd.DataFrame, y: pd.Series) -> bytes:
    # openpyxl directly, not pd.ExcelWriter. ExcelWriter's book has no active
    # sheet (it drops the default one), so `wb.remove(wb.active)` raises -- and
    # then saving raises again because a workbook must keep one visible sheet.
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)          # drop the default "Sheet"

    # ---------------------------------------------------- Adjustment grid
    ws = wb.create_sheet("Adjustment Grid")
    ws.cell(row=1, column=1, value="MARKET-DERIVED ADJUSTMENT GRID").font = TITLE_FONT
    r = _note(ws, 2, f"{meta['dataset_name']}  |  {meta['n_training_sales']:,} closed sales  "
                     f"|  {meta['date_min']} to {meta['date_max']}  |  model {meta['model_version']}")
    r = _note(ws, r, "Each figure is the average change in sale price from one additional unit "
                     "of the characteristic, holding everything else constant. Where the 95% "
                     "interval spans zero, the sample has not shown the effect differs from no "
                     "effect -- do not carry that line into a grid.")
    cols = ["Unit", "Adjustment", "CI_Lower_95", "CI_Upper_95",
            "Share_Responsive", "OLS_Crosscheck", "Use_In_Grid"]
    cols = [c for c in cols if c in grid.columns]
    _write(ws, grid[cols], start_row=r + 1, formats={
        "Adjustment": MONEY, "CI_Lower_95": MONEY, "CI_Upper_95": MONEY,
        "OLS_Crosscheck": MONEY, "Share_Responsive": '0%'})

    # -------------------------------------------- Percentage adjustments
    ws = wb.create_sheet("Percent Adjustments")
    ws.cell(row=1, column=1, value="PERCENTAGE ADJUSTMENTS").font = TITLE_FONT
    r = _note(ws, 2, "From a second model trained on the natural log of price. A percentage "
                     "effect is usually more stable across price tiers than a fixed dollar "
                     "amount. Close agreement with the dollar grid is good evidence the "
                     "adjustment is real.")
    pcols = [c for c in ["Unit", "Percent_Effect", "Pct_CI_Lower", "Pct_CI_Upper",
                         "Dollar_At_Mean_Price"] if c in pct.columns]
    _write(ws, pct[pcols], start_row=r + 1, formats={
        "Percent_Effect": PCT, "Pct_CI_Lower": PCT, "Pct_CI_Upper": PCT,
        "Dollar_At_Mean_Price": MONEY})

    # ------------------------------------------------------- Location
    if len(loc):
        ws = wb.create_sheet("Location")
        ws.cell(row=1, column=1, value="LOCATION ADJUSTMENTS").font = TITLE_FONT
        r = _note(ws, 2, f"Dollar difference relative to the baseline location "
                         f"({meta.get('baseline_location')}), everything else held constant.")
        _write(ws, loc, start_row=r + 1, formats={"Adjustment_vs_Baseline": MONEY})

    # ---------------------------------------------------- Importance
    ws = wb.create_sheet("Relative Importance")
    ws.cell(row=1, column=1, value="PERMUTATION IMPORTANCE").font = TITLE_FONT
    r = _note(ws, 2, "Drop in R-squared when each characteristic is randomly shuffled on the "
                     "held-out sales. Read this for RANKING; read the Adjustment Grid for the "
                     "dollar figure. A characteristic can rank high on one and low on the other.")
    _write(ws, importance, start_row=r + 1, formats={"Importance": DEC4})

    # ------------------------------------------------ Model comparison
    ws = wb.create_sheet("Model Comparison")
    ws.cell(row=1, column=1, value="XGBOOST vs RANDOM FOREST").font = TITLE_FONT
    r = _note(ws, 2, "Both evaluated on the same held-out later period. R-squared is NOT "
                     "comparable to another dataset's -- it depends on that sample's price "
                     "dispersion. Use MAE as a percent of mean price to compare markets.")
    c = meta["model_comparison"]
    cmp_df = pd.DataFrame({
        "Metric": ["R-squared (test)", "MAE (test)", "RMSE (test)", "MAE as % of mean price"],
        "XGBoost": [c["xgboost"]["r2"], c["xgboost"]["mae"],
                    c["xgboost"]["rmse"], c["xgboost"]["mae_pct"]],
        "Random Forest": [c["random_forest"]["r2"], c["random_forest"]["mae"],
                          c["random_forest"]["rmse"], c["random_forest"]["mae_pct"]],
    })
    r2 = _write(ws, cmp_df, start_row=r + 1)
    for row_i, fmt in zip(range(r + 2, r + 6), [DEC4, MONEY, MONEY, '0.00"%"']):
        for col_i in (2, 3):
            ws.cell(row=row_i, column=col_i).number_format = fmt
    _note(ws, r2, f"Selection: {c['verdict']}")

    # ------------------------------------------------ Sample statistics
    ws = wb.create_sheet("Sample Statistics")
    ws.cell(row=1, column=1, value="MODELING VARIABLES").font = TITLE_FONT
    stats = X.describe().T.reset_index().rename(columns={"index": "Feature"})
    stats = stats.round(2)
    _write(ws, stats, start_row=3, formats={c: NUM for c in stats.columns[1:]})

    # ------------------------------------------------------------ Basis
    ws = wb.create_sheet("Basis")
    ws.cell(row=1, column=1, value="BASIS OF THE ANALYSIS").font = TITLE_FONT
    d = meta["data_diagnostics"]
    basis = pd.DataFrame({
        "Item": [
            "Dataset", "Source file", "Model version", "Trained (UTC)",
            "Closed sales analyzed", "Sale date range", "Time split date",
            "Mean sale price", "Median price per sq ft",
            "Price span (max/min)", "Coefficient of variation",
            "Correlation, price to GLA", "Range restriction detected",
            "Concessions netted", "Fields absent from export",
            "Features dropped (no variation)", "Baseline location",
        ],
        "Value": [
            meta["dataset_name"], meta["source_file"], meta["model_version"],
            meta["trained_at"], f"{meta['n_training_sales']:,}",
            f"{meta['date_min']} to {meta['date_max']}", meta["split_date"],
            f"${meta['mean_price']:,.0f}", f"${d['raw_median_ppsf']:,.2f}",
            f"{d['price_range_ratio']:.2f}x", f"{d['price_cv']:.3f}",
            f"{d['gla_corr']:.3f}", "YES" if d["flags"] else "no",
            "yes" if meta["concessions_netted"] else "no",
            ", ".join(meta["missing_fields"]) or "none",
            ", ".join(meta["dropped_features"]) or "none",
            meta.get("baseline_location", "-"),
        ],
    })
    r = _write(ws, basis, start_row=3)

    for flag in d["flags"]:
        r = _note(ws, r, f"RANGE RESTRICTION: {flag}")
    if d["flags"]:
        r = _note(ws, r, "Every adjustment in this workbook is biased toward zero when the "
                         "sample is restricted on price. No modelling technique repairs this. "
                         "The fix is a wider MLS export.")
    r += 1
    for line in [
        "HOW THESE NUMBERS WERE PRODUCED",
        "Adjustments are counterfactual marginal effects from a monotone-constrained XGBoost "
        "ensemble: every closed sale is re-predicted with one more unit of the characteristic, "
        "and the differences are averaged. Intervals are percentile bootstraps over sales.",
        "There are no formulas in this workbook. A gradient-boosted marginal effect cannot be "
        "expressed as a spreadsheet formula, so these are recorded values, not calculations. "
        "To reproduce them, re-run the pipeline against the source file named above.",
        "Condition, quality of finish, updates, view, and functional utility are not in an MLS "
        "export. They are often the largest real difference between two comparables and are "
        "not in any of these numbers.",
        "This is a statistical support tool, not a value opinion, and does not satisfy USPAP "
        "on its own.",
    ]:
        r = _note(ws, r, line)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
